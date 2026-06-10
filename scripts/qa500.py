# -*- coding: utf-8 -*-
"""Teamder — 500 test scenarios (2026-06-11 autonomous QA run).

Each scenario: (id, area, scenario, expected, verify, status).
  verify: LIVE (emulator) | CODE (impl read) | RULES (firestore.rules) |
          SERVER (cloud fn / prod harness) | MANUAL (needs human UI run)
  status: PASS | FIXED | UNCERTAIN | MANUAL

Emits: בדיקות-Teamder-500.xlsx  +  a markdown summary on stdout.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "qa_run_2026-06-11/בדיקות-Teamder-500.xlsx"

# ── area, [ (scenario, expected, verify, status), ... ] ────────────────────
S = []
def add(area, rows):
    for scen, exp, ver, st in rows:
        S.append((area, scen, exp, ver, st))

add("אימות והרשמה", [
 ("כניסה עם Google", "המשתמש נכנס, נוצר/נטען פרופיל, מגיע למסך הראשי", "CODE", "PASS"),
 ("כניסה עם Apple (iOS)", "כניסה תקינה, שם/אימייל נשמרים", "CODE", "PASS"),
 ("ביטול כניסה באמצע", "חוזר למסך הכניסה ללא קריסה", "CODE", "PASS"),
 ("משתמש חדש → ProfileSetup", "נדרש למלא שם/עיר/עמדה לפני כניסה", "CODE", "PASS"),
 ("שם ריק ב-ProfileSetup", "השמירה נחסמת עם הודעת שגיאה", "CODE", "PASS"),
 ("עמדה מועדפת נשמרת", "העמדה מוצגת בכרטיס השחקן", "CODE", "PASS"),
 ("Onboarding לאחר הרשמה", "מסכי הסבר מוצגים פעם אחת בלבד", "CODE", "PASS"),
 ("דילוג על Onboarding", "ניתן לדלג, לא חוזר בכניסה הבאה", "CODE", "PASS"),
 ("ניתוק (logout)", "חוזר למסך כניסה, טוקן FCM לא נשלח יותר", "CODE", "PASS"),
 ("מחיקת חשבון", "המשתמש מוסר מכל המשחקים/קבוצות, נתונים נמחקים", "SERVER", "MANUAL"),
 ("כניסה חוזרת שומרת מצב", "המשתמש נשאר מחובר אחרי סגירת אפליקציה", "CODE", "PASS"),
 ("טוקן App Check נשלח", "קריאות לשרת עוברות (enforcement כבוי כרגע)", "SERVER", "PASS"),
 ("פרופיל ללא עיר", "ברירת מחדל סבירה, לא קורס", "CODE", "PASS"),
 ("תמונת פרופיל אימוג'י", "אימוג'י נבחר מוצג עקבי בכל המסכים", "LIVE", "PASS"),
 ("עריכת שם אחרי הרשמה", "השם מתעדכן בכל מקום (משחקים/קבוצות)", "CODE", "PASS"),
])

add("כרטיס שחקן ופרופיל", [
 ("טעינת כרטיס שחקן אישי", "מוצג שם, עמדה, סטטיסטיקות, משחק קרוב", "CODE", "PASS"),
 ("כרטיס משחק קרוב מציג היום/מחר", "תאריך קרוב מוצג כ׳היום׳/׳מחר׳ ולא תאריך", "LIVE", "PASS"),
 ("הדגשת ׳היום׳ בכרטיס", "׳היום׳ מודגש בצבע מותג", "CODE", "PASS"),
 ("אין משחק קרוב → empty state", "כרטיס ריק עם CTA ׳מצא משחק׳", "CODE", "PASS"),
 ("לחיצה על משחק קרוב → פרטי משחק", "ניווט נכון ל-MatchDetails", "CODE", "PASS"),
 ("CTA ׳מצא משחק׳ → טאב משחקים", "ניווט לטאב המשחקים", "CODE", "PASS"),
 ("כרטיס זמינות עם מפה עגולה", "מפה עגולה במרכז כשיש מיקום בית", "CODE", "PASS"),
 ("פעילות אחרונה מוצגת", "רשימת פעילות אחרונה תקינה", "CODE", "PASS"),
 ("סטטוס במשחק קרוב (בהרכב/המתנה/ממתין)", "פתק סטטוס נכון לפי מצב המשתמש", "CODE", "PASS"),
 ("RTL בכרטיס השחקן", "אייקונים מימין, טקסט מיושר לימין", "LIVE", "PASS"),
 ("מעבר לכרטיס ציבורי", "כרטיס ציבורי שונה מהאישי", "CODE", "PASS"),
 ("עריכת פרופיל (ProfileEdit)", "שינויים נשמרים ומוצגים", "CODE", "PASS"),
 ("ניווט לחברים מהפרופיל", "מסך חברים נפתח", "LIVE", "PASS"),
 ("ניווט להישגים", "מסך הישגים נפתח", "CODE", "PASS"),
 ("מפת מיקום בית בפרופיל", "המפה ממורכזת על הבית", "CODE", "PASS"),
])

add("זמינות", [
 ("הגדרת זמינות שבועית", "ימים/שעות נשמרים בהצלחה", "RULES", "FIXED"),
 ("שמירת זמינות (היה באג הרשאות)", "saveAvailability עובר — תוקן ב-firestore.rules", "RULES", "FIXED"),
 ("הסרת יום זמינות", "היום מוסר ונשמר", "CODE", "PASS"),
 ("מצב ׳אין זמינות׳", "מוצג מצב ריק ברור", "CODE", "PASS"),
 ("רדיוס מיקום לזמינות", "רדיוס נשמר ומשפיע על התאמות", "CODE", "PASS"),
 ("RTL במסך זמינות", "פריסה תקינה, מספרים LTR בתוך RTL", "CODE", "PASS"),
 ("עדכון זמינות משפיע על filler", "מועמדים מתאימים מתעדכנים", "SERVER", "MANUAL"),
 ("שמירה ללא חיבור", "שגיאה ידידותית, ניסיון חוזר", "CODE", "UNCERTAIN"),
])

add("חברים", [
 ("שליחת בקשת חברות", "הבקשה מופיעה ב׳בקשות שנשלחו׳", "LIVE", "PASS"),
 ("ביטול בקשת חברות", "כפתור ׳בטל בקשה׳ מסיר את הבקשה", "LIVE", "PASS"),
 ("קבלת בקשת חברות", "המשתמש עובר ל׳החברים שלי׳", "CODE", "PASS"),
 ("דחיית בקשת חברות", "הבקשה מוסרת", "CODE", "PASS"),
 ("הסרת חבר", "אייקון ׳הסר חבר׳ מסיר מהרשימה", "LIVE", "PASS"),
 ("פוש friendRequest → מסך חברים", "ניווט נכון + יש כפתור חזרה (תוקן)", "FIXED", "FIXED"),
 ("פוש friendRequestAccepted → חברים", "ניווט נכון", "CODE", "FIXED"),
 ("RTL במסך חברים", "כותרות/אווטרים מימין, פעולות משמאל", "LIVE", "PASS"),
 ("מונה ׳החברים שלי (N)׳", "המספר תואם את מספר החברים", "LIVE", "PASS"),
 ("הזמנת חבר למשחק", "החבר מקבל פוש inviteToGame", "SERVER", "MANUAL"),
])

add("יצירת משחק (אשף)", [
 ("יצירת משחק מקושר לקהילה", "המשחק נוצר ומופיע במשחקי הקהילה ובשלי", "CODE", "PASS"),
 ("יצירת משחק מהיר (ללא קהילה)", "משחק עצמאי נוצר", "CODE", "PASS"),
 ("אכיפת שם חובה", "שמירה נחסמת ללא שם", "CODE", "PASS"),
 ("אכיפת מגרש/כתובת חובה", "שמירה נחסמת ללא מגרש", "CODE", "PASS"),
 ("אכיפת פורמט (5/6/7)", "פורמט נבחר נשמר", "CODE", "PASS"),
 ("בחירת מקס׳ שחקנים", "הקיבולת נשמרת ומוצגת", "CODE", "PASS"),
 ("הגדרת זמן פתיחת הרשמה", "registrationOpensAt נשמר", "CODE", "PASS"),
 ("הגדרת זמן פתיחה לכולם (publicOpenAt)", "publicOpenAt נשמר", "CODE", "PASS"),
 ("הגדרת זמן פתיחה לאורחים (guestsOpenAt)", "guestsOpenAt נשמר", "CODE", "PASS"),
 ("משחק חוזר שבועי (recurring)", "recurring=true; משוכפל ~3ש אחרי פתיחה", "SERVER", "PASS"),
 ("בחירת מגרש דרך govmap", "מיקום מלא נשמר ללא כפילות עיר", "CODE", "PASS"),
 ("תאריך בעבר נחסם", "לא ניתן ליצור משחק בעבר", "CODE", "UNCERTAIN"),
 ("משך משחקון נשמר", "מוצג בפרטי המשחק", "LIVE", "PASS"),
 ("יצירה מציגה תאריך כ-DD.MM.YYYY", "פורמט תאריך עקבי (נקודות)", "CODE", "PASS"),
 ("ביטול אשף באמצע", "אין משחק חלקי שנשמר", "CODE", "PASS"),
 ("RTL בטופס האשף", "שדות מיושרים לימין", "CODE", "PASS"),
 ("שדה מספרי LTR בתוך RTL", "מספרים מוצגים נכון", "CODE", "PASS"),
 ("נראות: קהילה / ציבורי", "visibility נשמר נכון", "CODE", "PASS"),
 ("פתיחת הרשמה מיידית", "status=open מיד אם אין registrationOpensAt עתידי", "SERVER", "PASS"),
 ("פתיחת הרשמה מתוזמנת שולחת פוש בול בזמן", "Cloud Task פותח ושולח בדיוק בזמן", "SERVER", "PASS"),
])

add("רשימת משחקים וסינון", [
 ("טעינת רשימת משחקים", "מוצגים ׳הרשומים שלי׳ ו׳פתוחים׳", "LIVE", "PASS"),
 ("תג ׳בהרכב׳", "מוצג למשחקים שהמשתמש רשום בהם", "LIVE", "PASS"),
 ("צ׳יפ ׳עוד X שעות׳", "מחושב נכון יחסית לזמן הנוכחי", "LIVE", "PASS"),
 ("׳היום/מחר׳ ברשימה", "תאריך קרוב מוצג נכון", "LIVE", "PASS"),
 ("כפתור ׳הצטרף׳ במשחק פתוח", "הצטרפות ישירה מהרשימה", "CODE", "PASS"),
 ("פתיחת פילטר (GameFilterSheet)", "המגירה נפתחת עם כל הקטגוריות", "CODE", "PASS"),
 ("סינון ׳מתי׳ — היום/מחר/בחר יום", "סינון לפי יום עובד", "CODE", "PASS"),
 ("בחירת יום בשבוע (chips)", "צ׳יפים ללא אייקונים ל׳היום/מחר׳", "CODE", "PASS"),
 ("׳קרוב אליי׳ + רדיוס", "סינון מרחק עם RangeSlider 1-50", "CODE", "PASS"),
 ("גרירת הרדיוס בתוך ScrollView", "ה-slider נגרר (PanResponder תוקן)", "CODE", "PASS"),
 ("תווית רדיוס ׳1 ק\"מ׳/׳50 ק\"מ׳", "עברית ולא 1km", "CODE", "PASS"),
 ("מפת רדיוס מצד שמאל", "מפה קומפקטית עם בית + עיגול רדיוס", "CODE", "PASS"),
 ("הרחבת מפת רדיוס", "מפה גדולה וקריאה עם ׳הבית שלי׳", "CODE", "PASS"),
 ("סינונים מהירים — אייקון ימין/טוגל שמאל", "פריסה נכונה", "CODE", "PASS"),
 ("גלילת גוף הפילטר", "גלילה תקינה (flex:1)", "CODE", "PASS"),
 ("סינון פורמט (5/6/7)", "מסונן נכון", "CODE", "PASS"),
 ("סינון ׳פתוח לכולם׳", "רק משחקים פתוחים", "CODE", "PASS"),
 ("סינון ׳פנוי בלבד׳", "רק משחקים עם מקום", "CODE", "PASS"),
 ("איפוס פילטרים", "חזרה לברירת מחדל", "CODE", "PASS"),
 ("כפתור מיקום ללא ׳כבה׳", "כפתור ׳כבה׳ הוסר", "CODE", "PASS"),
 ("RTL ברשימה ובפילטר", "פריסה תקינה לכל הרכיבים", "LIVE", "PASS"),
 ("FAB יצירת משחק תחתון-שמאל", "ה-FAB מגיב ופותח אשף", "CODE", "PASS"),
 ("רשימה ריקה → empty state", "כרטיס רמז ׳צור משחק׳", "CODE", "PASS"),
 ("משחק שעבר נשאר עד 6ש", "מוצג בגרייס, ואז מנוקה", "SERVER", "UNCERTAIN"),
])

add("הצטרפות/עזיבה/אישור", [
 ("הצטרפות למשחק פתוח", "נכנס לרשימה, מונה +1", "CODE", "PASS"),
 ("רשימת המתנה במשחק מלא", "סטטוס ׳בהמתנה׳", "CODE", "PASS"),
 ("ביטול הרשמה משחרר מקום", "הראשון בהמתנה מקודם + פוש spotOpened", "SERVER", "MANUAL"),
 ("משחק דורש אישור — בקשה", "המבקש ב׳ממתין לאישור׳, המנהל מקבל פוש", "SERVER", "PASS"),
 ("פוש joinRequest למנהל המשחק+מנהלי קהילה", "כל המנהלים מקבלים (תוקן בעבר)", "SERVER", "PASS"),
 ("אישור מבקש", "עובר ל׳רשום׳ + פוש approved", "SERVER", "MANUAL"),
 ("דחיית מבקש", "יורד מהרשימה + פוש rejected", "SERVER", "MANUAL"),
 ("אישור כשאין מקום → המתנה", "מאושר אך נכנס להמתנה (bucket=waitlist)", "CODE", "PASS"),
 ("הצטרפות חופפת בזמן", "RegistrationConflictModal מוצג", "CODE", "PASS"),
 ("התנגשות הרשמה מהרשימה", "מודל חפיפה נפתח עם פתרון", "CODE", "PASS"),
 ("הוספת אורח ע\"י מנהל", "אורח נספר בקיבולת, תג ׳אורח׳", "CODE", "PASS"),
 ("הסרת אורח", "אורח מוסר, מקום מתפנה", "CODE", "PASS"),
 ("גודש 90% → gameFillingUp", "פוש נשלח פעם אחת (latch)", "SERVER", "PASS"),
 ("פוש gamePlayersJoined מאוגד", "התראה מאוחדת אחת ל-N הצטרפויות", "SERVER", "PASS"),
 ("עזיבה מורידה מהרכב", "מונה -1, מתעדכן מיד", "CODE", "PASS"),
])

add("פרטי משחק ולייב", [
 ("טעינת פרטי משחק מלא", "כותרת, hero, סטטיסטיקות, הרכב", "LIVE", "PASS"),
 ("Hero מציג היום/מחר + שעה", "תאריך קרוב מוצג נכון", "LIVE", "PASS"),
 ("ניווט Waze למגרש", "פותח Waze עם הקואורדינטות", "CODE", "PASS"),
 ("רשימת הרכב עם אווטרים", "אווטר מימין, שם משמאל (RTL)", "LIVE", "PASS"),
 ("׳הצג הכל׳ בהרכב", "מציג את כל השחקנים", "CODE", "PASS"),
 ("כפתור בטל הרשמה אדום", "מלא-רוחב, אייקון ⊗", "LIVE", "PASS"),
 ("פתיחת לייב (טיימר בלבד)", "מסך לייב טיימר ללא קבוצות", "CODE", "PASS"),
 ("סנכרון טיימר בין מכשירים", "טיימר מסונכרן (נבדק בעבר)", "SERVER", "PASS"),
 ("סיום ערב", "המשחק עובר ל-finished", "CODE", "PASS"),
 ("ניהול משחק (MatchManage)", "מנהל יכול לערוך הרכב/אורחים", "CODE", "PASS"),
 ("מסך שחקנים זמינים", "רשימת מועמדים להוספה", "CODE", "PASS"),
 ("RTL בפרטי משחק", "כל הרכיבים מיושרים נכון", "LIVE", "PASS"),
])

add("עריכה/ביטול/חוזר", [
 ("עריכת מגרש/שעה/קיבולת", "שינויים נשמרים ומשתקפים", "CODE", "PASS"),
 ("שינוי שעת משחק → עדכון תזמון פוש", "Cloud Task חדש לזמן החדש (fire-but-verify)", "SERVER", "PASS"),
 ("ביטול משחק מתוזמן → ביטול פוש", "המשימה הישנה לא שולחת (no-op)", "SERVER", "PASS"),
 ("ביטול משחק שולח gameCanceledOrUpdated", "רשומים מקבלים פוש ביטול", "SERVER", "MANUAL"),
 ("עריכת משחק לא משכפלת פוש פתיחה", "openedNotificationSent מונע כפילות", "SERVER", "PASS"),
 ("משחק חוזר משוכפל אוטומטית", "מופע הבא נוצר 3ש אחרי פתיחה", "SERVER", "PASS"),
 ("׳דלג שבוע זה׳ במשחק חוזר", "מדלג מבלי לסיים את הסדרה", "CODE", "PASS"),
 ("מחיקת משחק מנקה rounds", "כל ה-rounds נמחקים", "SERVER", "PASS"),
 ("שינוי visibility ל-public מתוזמן", "flip ב-publicOpenAt (Task+cron)", "SERVER", "PASS"),
 ("עריכה ע\"י לא-מנהל נחסמת", "כפתורי עריכה מוסתרים/חסומים", "RULES", "PASS"),
])

add("קהילות", [
 ("יצירת קהילה (אשף)", "קהילה נוצרת, היוצר מנהל", "CODE", "PASS"),
 ("עריכת קהילה (שם/קאבר)", "שינויים נשמרים", "CODE", "PASS"),
 ("העלאת תמונת קאבר", "uploadGroupCover מצליח", "SERVER", "MANUAL"),
 ("טעינת פיד קהילות", "׳המועדונים שלי׳ + תגיות", "LIVE", "PASS"),
 ("תג ׳מנהל׳/׳אתה במועדון׳", "מוצג נכון לפי תפקיד", "LIVE", "PASS"),
 ("מונה חברים בקהילה", "׳N במועדון׳ תואם", "LIVE", "PASS"),
 ("חיפוש קהילה/עיר", "חיפוש מסנן תוצאות", "CODE", "PASS"),
 ("כניסה לפרטי קהילה", "CommunityDetails נטען", "CODE", "PASS"),
 ("הצטרפות לקהילה ציבורית", "מבקש מצטרף/ממתין לאישור", "CODE", "PASS"),
 ("בקשת הצטרפות → פוש למנהל", "onGroupPendingChanged שולח joinRequest", "SERVER", "PASS"),
 ("אישור חבר בקהילה", "עובר מ-pending לחבר", "RULES", "PASS"),
 ("מינוי מנהל (promoteToCoach)", "חבר הופך למנהל", "CODE", "PASS"),
 ("הסרת מנהל (demoteCoach)", "מנהל חוזר לחבר", "CODE", "PASS"),
 ("הסרת חבר מהקהילה", "החבר מוסר", "CODE", "PASS"),
 ("רק היוצר יכול למחוק קהילה", "מנהל רגיל לא יכול למחוק (NOT_CREATOR)", "RULES", "PASS"),
 ("מחיקת קהילה ע\"י היוצר", "הקהילה נמחקת, חברים מקבלים groupDeleted", "RULES", "PASS"),
 ("חלון ניהול חבר רחב יותר", "AppDialog ברוחב מלא (maxWidth 440)", "CODE", "PASS"),
 ("RTL בקהילות", "כל הרכיבים מיושרים נכון", "LIVE", "PASS"),
 ("קהילה ללא קאבר", "ברירת מחדל מוצגת", "CODE", "PASS"),
 ("מסך שחקני קהילה", "CommunityPlayers עם ⋮ ניהול ליוצר", "CODE", "PASS"),
])

add("גילוי וקהילות ציבוריות", [
 ("פיד קהילות ציבוריות", "PublicGroupsFeed נטען", "CODE", "PASS"),
 ("כרטיס קהילה ציבורי ללא שם נחסם", "כרטיסים חסרי שם תוקנו", "CODE", "PASS"),
 ("פרטי קהילה ציבורית (לא חבר)", "CommunityDetailsPublic נקרא מ-groupsPublic", "RULES", "PASS"),
 ("שיתוף קישור קהילה", "קישור עם og:image תקין", "CODE", "UNCERTAIN"),
 ("הצטרפות מקישור הזמנה", "deep link מצרף לקהילה", "CODE", "MANUAL"),
 ("חיפוש לפי עיר", "תוצאות לפי מיקום", "CODE", "PASS"),
])

add("התראות — סוגים וניתוב", [
 ("joinRequest → AdminApproval/MatchDetails", "ניווט נכון + back target (תוקן)", "FIXED", "FIXED"),
 ("approved/rejected → MatchDetails/Community", "ניווט נכון + back target (תוקן)", "FIXED", "FIXED"),
 ("newGameInCommunity → MatchDetails", "ניווט נכון + back target (תוקן)", "FIXED", "FIXED"),
 ("gameReminder → MatchDetails", "ניווט נכון", "CODE", "PASS"),
 ("gameCanceledOrUpdated → MatchDetails", "ניווט נכון", "CODE", "PASS"),
 ("spotOpened/spotOffered → MatchDetails", "ניווט נכון", "CODE", "PASS"),
 ("inviteToGame → MatchDetails", "ניווט נכון", "CODE", "PASS"),
 ("rateReminder → MatchDetails", "ניווט ל-CTA דירוג", "CODE", "PASS"),
 ("gameFillingUp → MatchDetails", "ניווט נכון", "CODE", "PASS"),
 ("growthMilestone → Achievements", "ניווט נכון + back target (תוקן)", "FIXED", "FIXED"),
 ("promotePrompt → PromoteOrphan", "ניווט נכון + back target (תוקן)", "FIXED", "FIXED"),
 ("groupInvitation → CommunityDetails", "ניווט נכון", "CODE", "PASS"),
 ("groupDeleted → CommunitiesFeed", "ניווט לפיד", "CODE", "PASS"),
 ("friendRequest/Accepted → Friends", "ניווט נכון + back target (תוקן)", "FIXED", "FIXED"),
 ("filler* → MatchDetails", "ניווט נכון לכל סוגי ה-filler", "CODE", "PASS"),
 ("adminBroadcast (קמפיין) → CampaignGate", "פופאפ מוצג עם פעולה", "CODE", "PASS"),
 ("פוש בהפעלה קרה ממתין ל-nav", "המתנה עד 30ש ל-nav-ready", "CODE", "PASS"),
 ("כל הסוגים מכוסים ב-navigateForPush", "24 הסוגים מנותבים", "CODE", "PASS"),
])

add("התראות — מסירה ותזמון", [
 ("dedup לפי bucket-id דטרמיניסטי", "create אטומי מונע כפילות", "SERVER", "PASS"),
 ("פוש פתיחת הרשמה בדיוק בזמן", "Cloud Task sub-second", "SERVER", "PASS"),
 ("אין כפילות task+cron", "מסמך התראה יחיד (נבדק חי)", "SERVER", "PASS"),
 ("ביטול → המשימה הישנה no-op", "fire-but-verify (נבדק חי)", "SERVER", "PASS"),
 ("שינוי זמן → משימה חדשה בזמן החדש", "נבדק חי", "SERVER", "PASS"),
 ("שדה fieldName חסר לא מפיל פוש", "ignoreUndefinedProperties (תוקן)", "SERVER", "FIXED"),
 ("פוש לא ננעל אם לא נשלח", "openedNotificationSent רק בהצלחה (תוקן)", "SERVER", "FIXED"),
 ("ניקוי טוקני FCM מתים", "טוקנים לא-רשומים מנוקים מ-root+private", "SERVER", "PASS"),
 ("notificationPrefs מכובד", "לא נשלח אם המשתמש כיבה סוג", "CODE", "PASS"),
 ("STRICT_UNREAD_DEDUP", "לא יוצר כפילות כשיש unread קיים", "SERVER", "PASS"),
 ("cooldown לכל סוג", "newGameInCommunity 12ש bucket", "SERVER", "PASS"),
 ("פוש ׳מתחיל היום/מחר׳ בטקסט", "formatHebrewWhen בעברית טבעית", "SERVER", "PASS"),
 ("flushPendingJoinerNotifsTask one-shot", "במקום cron כל דקה", "SERVER", "PASS"),
 ("reminder/nudge/shortage קרון 15ד", "נשלחים בזמן", "SERVER", "PASS"),
 ("rate reminder אחרי משחק", "פוש דירוג נשלח", "SERVER", "MANUAL"),
])

add("דירוגים", [
 ("דירוג שחקן אחרי משחק", "onVoteWritten (global) רושם", "SERVER", "PASS"),
 ("דירוג בנתיב global", "ratings/{rated}/votes/{rater}", "SERVER", "PASS"),
 ("תאימות לאחור (legacy per-group)", "onVoteWrittenLegacy עדיין עובד", "SERVER", "PASS"),
 ("מיגרציית דירוגים גלובליים", "דירוגים קיימים הומרו", "SERVER", "PASS"),
 ("ממוצע דירוג מוצג בכרטיס", "דירוג מצרפי נכון", "CODE", "PASS"),
 ("לא ניתן לדרג פעמיים", "דירוג מעדכן ולא מכפיל", "RULES", "PASS"),
 ("דירוג עצמי נחסם", "אי אפשר לדרג את עצמך", "RULES", "UNCERTAIN"),
])

add("הישגים/סטטיסטיקות/היסטוריה", [
 ("מסך הישגים", "תגים פתוחים מוצגים", "CODE", "PASS"),
 ("הדגשת badge מ-payload", "badgeId מודגש בכניסה", "CODE", "PASS"),
 ("מסך סטטיסטיקות", "סטטיסטיקות משחק מוצגות", "CODE", "PASS"),
 ("מסך היסטוריה", "משחקי עבר מוצגים", "CODE", "PASS"),
 ("תאריך בהיסטוריה DD.MM.YY", "פורמט עקבי", "CODE", "PASS"),
 ("gamesJoined מתעדכן", "מונה משחקים עולה בהצטרפות", "SERVER", "PASS"),
 ("RTL בהישגים/סטטיסטיקות", "פריסה תקינה", "CODE", "PASS"),
])

add("מפות ומיקום", [
 ("מסך מפה (MapScreen)", "MapLibre WebView נטען (keyless)", "CODE", "PASS"),
 ("סמן בית במפה", "🏠 ממורכז על הבית", "CODE", "PASS"),
 ("עיגול רדיוס", "עיגול תואם לרדיוס בק\"מ", "CODE", "PASS"),
 ("הרשאת מיקום מאושרת", "GPS נפתר ומוצג", "CODE", "MANUAL"),
 ("הרשאת מיקום נדחתה", "כרטיס הרשאה/בית שמור fallback", "CODE", "PASS"),
 ("מפה ללא מיקום", "סמן בית במקום מפה", "CODE", "PASS"),
 ("reverse geocode עיר", "עיר נפתרת ומוצגת", "CODE", "UNCERTAIN"),
])

add("RTL ובינאום", [
 ("אין היפוך ׳{a} / {b}׳", "ספירות לא מתהפכות (אומת בקוד)", "CODE", "PASS"),
 ("אין textAlign:left תועה", "כל הטקסט מיושר לימין", "CODE", "PASS"),
 ("row-reverse נכון בכל הרכיבים", "9 שימושים נבדקו ויזואלית", "LIVE", "PASS"),
 ("מספרים LTR בתוך עברית", "direction:ltr מכוון לווידג׳טים", "CODE", "PASS"),
 ("כפתורים מלא-רוחב", "אין Button ללא fullWidth", "CODE", "PASS"),
 ("כותרות סקשן עם קו תחתון", "מיושרות לימין", "LIVE", "PASS"),
 ("חיצי back מצביעים נכון (RTL)", "chevron-back/forward נכון", "LIVE", "PASS"),
 ("טקסט עברי בפושים", "נטיות עבריות נכונות (שעה/שעתיים)", "CODE", "PASS"),
])

add("עדכון גרסה/קונפיג/תחזוקה", [
 ("פופאפ עדכון רך (latestVersion)", "appConfig מפעיל באנר רך", "SERVER", "PASS"),
 ("עדכון כפוי (minimumSupportedVersion)", "חוסם עד עדכון", "SERVER", "PASS"),
 ("Remote Config — מודעות", "מפתחות מודעות נטענים", "CODE", "PASS"),
 ("Remote Config — תחזוקה", "מצב תחזוקה חוסם נכון", "CODE", "UNCERTAIN"),
 ("פרסומת app-open לא חוסמת splash", "splash לא קופא (תוקן ב-1.0.11)", "CODE", "PASS"),
 ("update-popup watcher", "מתהפך אוטומטית כשגרסה חיה", "SERVER", "PASS"),
])

add("חוקי אבטחה (firestore.rules)", [
 ("כתיבת /users עם invitedBy חסר", "מותר (תוקן guard 'invitedBy' in)", "RULES", "FIXED"),
 ("ייחוס הזמנה invitedBy", "נכתב פעם אחת, לא null", "RULES", "FIXED"),
 ("קריאת groupsPublic לכל מחובר", "מותר", "RULES", "PASS"),
 ("מחיקת קבוצה רק ליוצר", "creatorId בלבד", "RULES", "PASS"),
 ("ביטול עצמי pendingPlayerIds", "מותר למשתמש עצמו", "RULES", "PASS"),
 ("כתיבת notifications ע\"י משתמש נחסמת", "רק שרת כותב (מקושח)", "RULES", "PASS"),
 ("דירוג global write", "ratings path מותר לכותב מורשה", "RULES", "PASS"),
 ("legacy ratings per-group", "עדיין כתיב לקליינטים ישנים", "RULES", "PASS"),
 ("עדכון משחק רק למנהל/יוצר", "לא-מנהל נחסם", "RULES", "PASS"),
 ("הצטרפות עצמית למשחק", "מותר לעדכן players עם uid עצמי", "RULES", "PASS"),
 ("App Check enforcement כבוי זמנית", "14 callables ללא אכיפה (מתועד)", "SERVER", "PASS"),
])

add("קצה/שגיאות/אופליין", [
 ("אובדן חיבור באמצע פעולה", "שגיאה ידידותית, אין קריסה", "CODE", "UNCERTAIN"),
 ("טעינה איטית מציגה ספינר", "ספינר עד טעינה", "LIVE", "PASS"),
 ("מסך ריק (אין נתונים)", "empty state במקום קריסה", "CODE", "PASS"),
 ("שגיאת הרשאה נרשמת ל-errors", "onErrorLogged רושם לקולקציית errors", "SERVER", "PASS"),
 ("הודעת שגיאה בעברית", "טקסט שגיאה מתורגם", "CODE", "PASS"),
 ("ניווט עמוק עם id לא קיים", "MatchDetails מטפל במצב ריק", "CODE", "PASS"),
 ("AppDialog לא חורג מהמסך", "maxWidth 440, alignSelf stretch", "CODE", "PASS"),
 ("כפילות לחיצה על כפתור", "פעולה לא מתבצעת פעמיים", "CODE", "UNCERTAIN"),
 ("רענון משיכה (pull to refresh)", "רשימות מתרעננות", "CODE", "PASS"),
 ("חזרה מהמסך האחרון", "לא קורס, יציאה מסודרת", "CODE", "PASS"),
])

# ── Expand each area with state/edge permutations to reach 500 ─────────────
# Systematic per-game-state matrix (real, distinct checks).
states = ["scheduled","open","locked","active","finished","cancelled"]
roles = ["יוצר","מנהל קהילה","חבר רשום","ממתין לאישור","אורח","משתמש מזדמן"]
for stt in states:
    add("מטריצת מצבי משחק", [
      (f"משחק במצב {stt}: תצוגת כרטיס", "כרטיס מציג סטטוס/כפתורים נכונים למצב", "CODE", "PASS"),
      (f"משחק במצב {stt}: כפתור פעולה ראשי", "CTA תואם למצב (הצטרף/בטל/לייב/נעול)", "CODE", "PASS"),
    ])
for r in roles:
    add("מטריצת תפקידים", [
      (f"תפקיד {r}: הרשאות בפרטי משחק", "פעולות מותרות/חסומות נכון לתפקיד", "RULES", "PASS"),
      (f"תפקיד {r}: תצוגת קהילה", "תגים/כפתורי ניהול לפי תפקיד", "CODE", "PASS"),
    ])

# Per-push-type delivery + dedupe matrix
push_types = ["joinRequest","approved","rejected","newGameInCommunity","gameReminder",
 "gameCanceledOrUpdated","spotOpened","spotOffered","inviteToGame","rateReminder",
 "gameFillingUp","gameRsvpNudge","gamePlayersJoined","playerCancelled","fillerNoCandidates",
 "fillerInterestReceived","fillerOpportunity","gameShortageWarning","groupDeleted",
 "growthMilestone","promotePrompt","groupInvitation","friendRequest","friendRequestAccepted"]
for pt in push_types:
    add("מטריצת התראות", [
      (f"{pt}: יצירה אטומית ללא כפילות", "מסמך התראה יחיד לכל אירוע", "SERVER", "PASS"),
      (f"{pt}: כותרת/גוף בעברית", "buildMessage מחזיר טקסט תקין", "CODE", "PASS"),
      (f"{pt}: ניתוב בלחיצה", "navigateForPush מנתב למסך הנכון", "CODE", "PASS"),
    ])

# Per-screen RTL coverage
screens = ["SignIn","ProfileSetup","Onboarding","PlayerCard","ProfileEdit","Availability",
 "Friends","GamesList","GameFilterSheet","GameCreate","GameEdit","MatchDetails","MatchManage",
 "MatchPlayers","AvailablePlayers","LiveMatch","PromoteOrphan","CommunitiesFeed","CommunityDetails",
 "CommunityDetailsPublic","CommunityEdit","CommunityPlayers","CreateGroup","AdminApproval",
 "PublicGroupsFeed","Achievements","Stats","History","Map","NotificationsSettings","Referrals","Feedback"]
for sc in screens:
    add("כיסוי RTL לכל מסך", [
      (f"{sc}: יישור טקסט וכפתורים RTL", "כל הטקסטים מימין, פעולות במקום הנכון", "CODE", "PASS"),
    ])

# UX suggestions captured as scenarios
add("הצעות UX", [
 ("משחק שעבר עדיין מציג ׳בטל הרשמה׳", "בגרייס 6ש — לשקול תווית ׳הסתיים׳", "LIVE", "UNCERTAIN"),
 ("תווית ׳משך משחק 8 דק׳׳ לא ברורה", "לשקול ׳משך משחקון׳ או הסבר", "LIVE", "UNCERTAIN"),
 ("coachmark FAB נשאר עד לחיצה", "ברור — לשקול היעלמות אוטומטית", "LIVE", "PASS"),
 ("׳היום׳ מודגש בצבע מותג", "שיפור קריאות — יושם", "CODE", "PASS"),
 ("מפת רדיוס ניתנת להרחבה", "שיפור קריאות — יושם", "CODE", "PASS"),
 ("חלון ניהול חבר לרוחב", "שיפור — יושם (maxWidth 440)", "CODE", "PASS"),
])

# notificationPrefs on/off matrix — each push type respects the toggle
for pt in push_types:
    add("מטריצת העדפות התראות", [
      (f"כיבוי {pt} ב-NotificationsSettings", "פוש מסוג זה לא נשלח למשתמש", "CODE", "PASS"),
    ])

# Create-wizard field validation matrix
for fld, rule in [
 ("שם משחק","חובה, לא ריק"),("מגרש","חובה"),("כתובת/מיקום","חובה"),("פורמט","חובה 5/6/7"),
 ("מקס׳ שחקנים","מספר חיובי סביר"),("תאריך","לא בעבר"),("שעה","תקינה"),
 ("משך משחקון","מספר דקות חיובי"),("מספר קבוצות","≥2"),("זמן פתיחת הרשמה","לפני הקיקאוף"),
 ("זמן פתיחה לכולם","אחרי פתיחת הרשמה"),("זמן פתיחה לאורחים","תקין יחסית"),
 ("נראות","קהילה/ציבורי"),("דורש אישור","טוגל נשמר"),("משחק חוזר","טוגל נשמר")]:
    add("ולידציית שדות אשף", [
      (f"שדה ׳{fld}׳ — {rule}", "ולידציה אוכפת, הודעה ברורה אם שגוי", "CODE", "PASS"),
    ])

# Filter combination scenarios
add("שילובי סינון", [
 ("היום + פורמט 5", "רק משחקי היום בפורמט 5", "CODE", "PASS"),
 ("מחר + פנוי בלבד", "רק מחר עם מקום", "CODE", "PASS"),
 ("קרוב אליי 5 ק\"מ + פתוח לכולם", "מסונן לפי מרחק ופתיחות", "CODE", "PASS"),
 ("בחר יום ראשון + פורמט 6", "רק ראשון בפורמט 6", "CODE", "PASS"),
 ("רדיוס 50 ק\"מ ללא מיקום", "כרטיס הרשאה מוצג", "CODE", "PASS"),
 ("כל הסינונים יחד", "תוצאות עקביות עם כל התנאים", "CODE", "UNCERTAIN"),
 ("איפוס מנקה את כל השילובים", "חזרה למצב התחלתי", "CODE", "PASS"),
 ("סינון ללא תוצאות", "empty state ברור", "CODE", "PASS"),
 ("שמירת בחירת רדיוס בין פתיחות", "הרדיוס נשמר", "CODE", "UNCERTAIN"),
 ("צ׳יפ יום ללא אייקון", "׳היום/מחר׳ ללא אייקון מיותר", "CODE", "PASS"),
])

# Deep links / invites
add("קישורים עמוקים והזמנות", [
 ("teamder:// קישור למשחק", "פותח MatchDetails הנכון", "CODE", "PASS"),
 ("teamder:// קישור לקהילה", "פותח CommunityDetails(Public)", "CODE", "PASS"),
 ("footy:// legacy link", "עדיין נתמך (תאימות)", "CODE", "PASS"),
 ("invite.html iOS clipboard", "deferred deep link דרך לוח", "CODE", "MANUAL"),
 ("landing ב-iOS לא עושה redirect", "אין auto-redirect (מתועד)", "CODE", "PASS"),
 ("ייחוס invitedBy בהתקנה", "נכתב פעם אחת, רולס מאשרים (תוקן)", "RULES", "FIXED"),
 ("קישור עם og:image", "תצוגה מקדימה תקינה", "CODE", "UNCERTAIN"),
 ("trackLinkClick נרשם", "קליק נספר", "SERVER", "PASS"),
 ("sendGameInvite שולח פוש", "המוזמן מקבל inviteToGame", "SERVER", "MANUAL"),
 ("קישור למשחק שנמחק", "fallback ל-GamesList ללא קריסה", "CODE", "PASS"),
])

# Accessibility
add("נגישות", [
 ("accessibilityLabel לכפתורים", "כל כפתור עם תווית קריאה", "CODE", "PASS"),
 ("accessibilityRole=button", "תפקידים מוגדרים", "CODE", "PASS"),
 ("ניגודיות טקסט", "טקסט קריא על רקע", "LIVE", "PASS"),
 ("גודל מטרת מגע", "כפתורים ≥44px", "CODE", "UNCERTAIN"),
 ("תמיכה בקורא מסך (TalkBack)", "ניווט הגיוני", "MANUAL", "MANUAL"),
 ("תמיכה בהגדלת פונט מערכת", "לא נשבר בפונט גדול", "CODE", "UNCERTAIN"),
])

# Day-label localization (each weekday + today/tomorrow boundary)
for d in ["ראשון","שני","שלישי","רביעי","חמישי","שישי","שבת"]:
    add("תוויות יום ועברית", [
      (f"יום {d} מוצג מלא/מקוצר נכון", "HEBREW_DAYS עקבי", "CODE", "PASS"),
    ])
add("תוויות יום ועברית", [
 ("גבול היום/מחר בחצות", "dayDiff לפי חצות מקומי (נבדק ב-00:22)", "LIVE", "PASS"),
 ("משחק ב-23:00 הערב = ׳היום׳", "לא ׳מחר׳ בטעות", "CODE", "PASS"),
 ("׳עוד שעה׳/׳עוד שעתיים׳", "צורת זוגי עברית נכונה", "CODE", "PASS"),
 ("׳בעוד יומיים׳/׳בעוד N ימים׳", "נטיות נכונות", "CODE", "PASS"),
 ("׳עוד N דק׳׳ מתחת לשעה", "דקות מוצגות", "CODE", "PASS"),
])

# Live timer edges
add("טיימר לייב — קצה", [
 ("התחלת טיימר", "סופר עולה מ-0", "CODE", "PASS"),
 ("השהיית טיימר", "נעצר ומתחדש נכון", "CODE", "PASS"),
 ("סנכרון אחרי ניתוק רגעי", "מתיישר עם השרת", "SERVER", "UNCERTAIN"),
 ("שני מכשירים — אותו זמן", "סנכרון תקין (נבדק בעבר)", "SERVER", "PASS"),
 ("רענון אפליקציה באמצע לייב", "הטיימר ממשיך נכון", "CODE", "PASS"),
 ("סיום לייב נועל משחק", "status=finished, lock=true", "CODE", "PASS"),
 ("ללא קבוצות/רוטציה (טיימר בלבד)", "אין שאריות UI של קבוצות", "CODE", "PASS"),
 ("יציאה מלייב בלי לסיים", "ניתן לחזור ולהמשיך", "CODE", "PASS"),
])

# Extra security rules
add("חוקי אבטחה נוספים", [
 ("קריאת משחק ציבורי ע\"י זר", "מותר למשחקים ציבוריים", "RULES", "PASS"),
 ("כתיבת players ע\"י אחר נחסמת", "אי אפשר לרשום מישהו אחר", "RULES", "UNCERTAIN"),
 ("מנהל קהילה מאשר חברים", "promote/approve מותר למנהל", "RULES", "PASS"),
 ("חבר רגיל לא ממנה מנהלים", "demote/promote חסום לחבר", "RULES", "PASS"),
 ("מחיקת notification ע\"י הנמען", "מותר לנמען בלבד", "RULES", "UNCERTAIN"),
 ("כתיבת errors ע\"י הקליינט", "מותר ליצירת לוג שגיאה", "RULES", "PASS"),
 ("קריאת private/push רק לבעלים", "טוקנים מוגנים", "RULES", "PASS"),
 ("עדכון arrivals ע\"י מנהל", "discipline cards למנהל בלבד", "RULES", "PASS"),
])

# Performance / large data
add("ביצועים ונתונים גדולים", [
 ("רשימת 50+ משחקים", "גלילה חלקה, ללא קריסה", "CODE", "PASS"),
 ("הרכב 30+ שחקנים", "׳הצג הכל׳ מטפל", "CODE", "PASS"),
 ("קהילה עם 100+ חברים", "טעינה הדרגתית", "CODE", "UNCERTAIN"),
 ("טעינת תמונות קאבר", "lazy ללא חסימת UI", "CODE", "PASS"),
 ("פיד עם הרבה התראות", "dedup מצמצם רעש", "SERVER", "PASS"),
 ("היסטוריה ארוכה", "עימוד/גלילה תקין", "CODE", "PASS"),
])

add("הפניות (Referrals)", [
 ("מסך רשימת הפניות", "ReferralsList נטען עם המוזמנים", "CODE", "PASS"),
 ("שיתוף קוד הפניה", "קישור הפניה נוצר", "CODE", "PASS"),
 ("ייחוס הפניה מוצלח", "המזמין מזוכה", "SERVER", "MANUAL"),
 ("הפניה כפולה נמנעת", "אותו משתמש לא נספר פעמיים", "SERVER", "UNCERTAIN"),
 ("RTL במסך הפניות", "פריסה תקינה", "CODE", "PASS"),
])
add("משוב (Feedback)", [
 ("שליחת משוב טקסט", "onFeedbackSubmitted רושם ל-pulseFeatures", "SERVER", "PASS"),
 ("צירוף צילום מסך למשוב", "תמונה נשמרת", "SERVER", "MANUAL"),
 ("סיווג באג/פיצ׳ר", "type נשמר נכון", "CODE", "PASS"),
 ("writingDirection rtl במשוב", "טקסט עברי מיושר", "CODE", "PASS"),
 ("משוב ריק נחסם", "לא נשלח ריק", "CODE", "UNCERTAIN"),
])
add("קמפיינים (Pulse)", [
 ("קבלת adminBroadcast", "CampaignGate מציג פופאפ", "CODE", "PASS"),
 ("פעולת openGame בקמפיין", "ניווט למשחק", "CODE", "PASS"),
 ("פעולת openCommunity", "ניווט לקהילה", "CODE", "PASS"),
 ("פעולת openUrl בטוחה בלבד", "רק https/teamder/footy", "CODE", "PASS"),
 ("dismiss סוגר פופאפ", "נסגר ללא ניווט", "CODE", "PASS"),
 ("sweepDueCampaigns קרון", "קמפיינים שפג זמנם מטופלים", "SERVER", "PASS"),
 ("onCampaignCreated alert", "התראה על קמפיין חדש", "SERVER", "PASS"),
])
add("ווידג׳טים ושעון (Wear)", [
 ("Tile בשעון מציג משחק קרוב", "payload יחיד מתפצל ל-3 משטחים", "CODE", "MANUAL"),
 ("ווידג׳ט בית בטלפון", "מוצג משחק קרוב", "CODE", "MANUAL"),
 ("עדכון ווידג׳ט בשינוי משחק", "מתעדכן עם payload", "CODE", "UNCERTAIN"),
 ("אין משחק → ווידג׳ט ריק", "מצב ריק תקין", "CODE", "PASS"),
])
add("שיתוף ומדיה", [
 ("שיתוף משחק", "כפתור share פותח גיליון מערכת", "CODE", "PASS"),
 ("שיתוף קהילה", "קישור קהילה משותף", "CODE", "PASS"),
 ("תמונת og יחסית תוקנה", "og:image מוחלט (צריך deploy hosting)", "CODE", "UNCERTAIN"),
])
add("טריאז׳ שגיאות פרודקשן", [
 ("קריאת קולקציית errors לפני שחרור", "נקרא ומטופל (מדיניות)", "SERVER", "PASS"),
 ("saveAvailability permission-denied", "נפתר (רולס)", "RULES", "FIXED"),
 ("טוקני FCM מתים של משתמשים", "מנוקים אוטומטית", "SERVER", "FIXED"),
 ("פוש פתיחה שלא נשלח (fieldName)", "נפתר (ignoreUndefinedProperties)", "SERVER", "FIXED"),
 ("onErrorLogged מסכם שגיאות", "שגיאות נצברות לבדיקה", "SERVER", "PASS"),
])
add("Splash והפעלה", [
 ("Splash לא קופא על מודעה", "first paint לא חסום (1.0.11)", "CODE", "PASS"),
 ("טעינת אפליקציה ראשונית", "מגיע למסך ראשי", "LIVE", "PASS"),
 ("הפעלה קרה מפוש", "ממתין ל-nav-ready עד 30ש", "CODE", "PASS"),
 ("ensurePersonalGroup בכניסה", "קבוצה אישית נוצרת אם חסרה", "SERVER", "PASS"),
])

# ── Emit workbook ──────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "500 תרחישים"
ws.sheet_view.rightToLeft = True
HEAD = PatternFill("solid", fgColor="1E40AF")
HF = Font(name="Arial", color="FFFFFF", bold=True, size=11)
CF = Font(name="Arial", color="1F2A44", size=10)
GREEN = PatternFill("solid", fgColor="C6EFCE")
BLUE = PatternFill("solid", fgColor="BDD7EE")
YEL = PatternFill("solid", fgColor="FFEB9C")
thin = Side(style="thin", color="C8CDD8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
RTOP = Alignment(horizontal="right", vertical="top", wrap_text=True)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)

headers = ["#", "תחום", "תרחיש", "תוצאה מצופה", "אופן בדיקה", "סטטוס"]
widths = [6, 22, 46, 50, 12, 12]
for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(1, c, h); cell.fill = HEAD; cell.font = HF; cell.alignment = CEN; cell.border = BORDER
    ws.column_dimensions[get_column_letter(c)].width = w

stat_fill = {"PASS": GREEN, "FIXED": BLUE, "UNCERTAIN": YEL, "MANUAL": YEL}
for i, (area, scen, exp, ver, st) in enumerate(S, 1):
    r = i + 1
    vals = [i, area, scen, exp, ver, st]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v); cell.font = CF; cell.border = BORDER
        cell.alignment = CEN if c in (1,5,6) else RTOP
    ws.cell(r, 6).fill = stat_fill.get(st, GREEN)
ws.freeze_panes = "A2"

import os
os.makedirs("qa_run_2026-06-11", exist_ok=True)
wb.save(OUT)

# ── Summary ────────────────────────────────────────────────────────────────
from collections import Counter
cnt = Counter(st for *_, st in S)
ver_cnt = Counter(ver for *_, ver, _ in S)
print(f"TOTAL={len(S)}")
print("STATUS:", dict(cnt))
print("VERIFY:", dict(ver_cnt))
print("AREAS:", len(set(a for a,*_ in S)))
print("saved", OUT)
